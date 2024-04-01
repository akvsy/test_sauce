from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait #ilgili driverı bekleten yapı
from selenium.webdriver.support import expected_conditions as ec #beklenen koşullar
from selenium.webdriver.common.action_chains import ActionChains 
import pytest
import openpyxl
#import json
from constants.globalConstants import *


class Test_Login:
    
    def setup_method(self):
        self.driver= webdriver.Chrome()
        self.driver.maximize_window()
        self.driver.get(BASE_URL)
    
    def teardown_method(self):
        self.driver.quit()

    def login_info(self,username,password):
        usernameInput = self.waitForElementVisible((By.ID, username_id))
        passwordInput = self.waitForElementVisible((By.ID, password_id))
        usernameInput.send_keys(username)
        passwordInput.send_keys(password)
        loginButton = self.waitForElementVisible((By.ID, login_button_id))
        loginButton.click()

    def waitForElementVisible(self,locator,timeout=5):
        return WebDriverWait(self.driver,timeout).until(ec.visibility_of_element_located(locator))

    def readInvalidDataFromExcel():
        excelFile= openpyxl.load_workbook("data/invalidLogin.xlsx")
        sheet= excelFile["Sayfa1"]
        rows= sheet.max_row
        data= []
        for i in range(2, rows+1):
            username= sheet.cell(i,1).value
            password= sheet.cell(i,2).value
            data.append((username, password))
        return data
        
    # def readInvalidDataFromJSON():
    #     with open("invalidLogin.json") as f:
    #         data= json.load(f)
    #     return data

    @pytest.mark.parametrize("username, password", readInvalidDataFromExcel())
    def test_invalid_login(self,username,password):
        self.login_info(username, password)
        errorMessage = self.waitForElementVisible((By.XPATH, errorMessage_xpath))
        assert errorMessage.text == invalid_errorMessage


    def test_empty_field(self):
        self.login_info("", "")
        errorMessage = self.waitForElementVisible((By.XPATH, errorMessage_xpath))
        assert errorMessage.text == emptyField_errorMessage

    def test_empty_password(self):
        self.login_info("standard_user", "")
        errorMessage = self.waitForElementVisible((By.XPATH, errorMessage_xpath))
        assert errorMessage.text == emptyPassword_errorMessage

    def test_locked_out(self):
        self.login_info("locked_out_user", "secret_sauce")
        errorMessage = self.waitForElementVisible((By.XPATH, errorMessage_xpath))
        assert errorMessage.text == lockedOut_errorMessage

    def test_valid_login(self):
        self.login_info("standard_user", "secret_sauce")
        assert LOGIN_URL in self.driver.current_url

        itemList= self.driver.find_elements(By.CLASS_NAME, item_class_name)
        assert len(itemList)==6

    def test_add_to_cart(self): #sepete ürün ekleme
        self.login_info("standard_user", "secret_sauce")
        sleep(2)
        self.driver.execute_script("window.scrollTo(0, 500)")
        addToCart= self.waitForElementVisible((By.XPATH, addToCard_xpath))
        addToCart.click()
        remove= self.waitForElementVisible((By.XPATH, remove_xpath))
        assert remove.is_displayed()
    
    def test_low_to_high(self): #fiyatları düşükten yükseğe sıralama
        self.login_info("standard_user", "secret_sauce")
        productSort= self.waitForElementVisible((By.XPATH, productSort_xpath))
        productSort.click()
        lowToHigh= self.waitForElementVisible((By.XPATH, lowToHigh_xpath))
        lowToHigh.click()
        priceList= WebDriverWait(self.driver,5).until(ec.visibility_of_any_elements_located((By.CLASS_NAME, itemPrice_class_name)))
        prices = [float(price.text.strip('$')) for price in priceList]
        sortedPrice= sorted(prices)
        assert sortedPrice == prices

